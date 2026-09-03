# -*- coding: utf-8 -*-
"""
Convierte el inventario exportado del sistema anterior (.xlsx) al CSV que
importa TANICHI en Inventario -> Importar CSV.

Uso:  py convertir-inventario.py <archivo.xlsx> [salida.csv]
"""
import sys, csv, collections, os
import openpyxl

ORIGEN  = sys.argv[1] if len(sys.argv) > 1 else 'inventario.xlsx'
DESTINO = sys.argv[2] if len(sys.argv) > 2 else 'catalogo_tanichi.csv'

# Columnas del archivo de origen (0-based)
COL = dict(categoria=1, nombre=2, descripcion=3, codigo=5,
           costo=6, mayoreo=7, precio=8, existencia=12, activo=14)

STOCK_MIN = 5          # el mismo valor por defecto que usa la app


def numero(v):
    """Convierte a float tolerando comas, espacios y celdas vacías."""
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(',', '').strip())
    except ValueError:
        return 0.0


def texto(v):
    return ('' if v is None else str(v)).strip()


wb = openpyxl.load_workbook(ORIGEN, data_only=True)
filas = list(wb[wb.sheetnames[0]].iter_rows(min_row=2, values_only=True))

# Un código repetido haría que un producto sobrescriba al otro al importar,
# porque la app empareja por código. Se detectan antes de escribir nada.
repetidos = {c for c, n in collections.Counter(
    texto(f[COL['codigo']]) for f in filas).items() if n > 1 and c}

productos, avisos = [], []
vistos = set()          # códigos repetidos ya asignados a un producto
sin_codigo = 0

for i, f in enumerate(filas, start=2):
    nombre = texto(f[COL['nombre']])
    if not nombre:
        avisos.append(f'fila {i}: sin nombre, se omite')
        continue

    codigo = texto(f[COL['codigo']])
    if codigo in repetidos:
        # La app empareja por código: dos filas con el mismo código se
        # fundirían en un solo producto. Se conserva en la primera y se deja
        # en blanco en las siguientes, que entran identificadas por nombre.
        if codigo in vistos:
            avisos.append(f'fila {i}: código "{codigo}" ya usado por otro producto -> "{nombre}" entra sin código')
            codigo = ''
        else:
            vistos.add(codigo)
    if not codigo:
        sin_codigo += 1

    precio = numero(f[COL['precio']])          # Precio Publico
    costo  = numero(f[COL['costo']])
    activo = texto(f[COL['activo']]) == '1'

    if precio <= 0 and activo:
        avisos.append(f'fila {i}: "{nombre}" está activo con precio {precio}')

    # Las recargas tienen su propio flujo en la app (botón Recarga), que además
    # descuenta el saldo de Mercado Pago y suma la comisión devuelta. Vendidas
    # como producto normal, el corte contaría el efectivo pero no el movimiento
    # de MP, y el cuadre saldría mal en cada venta. Entran desactivadas.
    if activo and any(t in nombre.lower() for t in ('recarga', 'tiempo aire')):
        activo = False
        avisos.append(f'fila {i}: "{nombre}" entra DESACTIVADO -> usa el botón Recarga del POS')

    productos.append({
        'codigo': codigo,
        'nombre': nombre,
        'categoria': texto(f[COL['categoria']]) or 'General',
        'precio': round(precio, 2),
        'costo': round(costo, 2),
        'existencia': round(numero(f[COL['existencia']]), 3),
        'minimo': STOCK_MIN,
        'controla_stock': 'si',
        'activo': 'si' if activo else 'no',
    })

CABECERA = ['codigo', 'nombre', 'categoria', 'precio', 'costo',
            'existencia', 'minimo', 'controla_stock', 'activo']

# utf-8-sig: el BOM hace que Excel abra los acentos bien, y la app lo ignora
with open(DESTINO, 'w', newline='', encoding='utf-8-sig') as fh:
    w = csv.DictWriter(fh, fieldnames=CABECERA, lineterminator='\r\n')
    w.writeheader()
    w.writerows(productos)

activos = sum(1 for p in productos if p['activo'] == 'si')
valor   = sum(p['precio'] * p['existencia'] for p in productos if p['activo'] == 'si')
costo   = sum(p['costo']  * p['existencia'] for p in productos if p['activo'] == 'si')

print(f'Archivo    : {os.path.abspath(DESTINO)}')
print(f'Productos  : {len(productos)}  ({activos} activos, {len(productos)-activos} inactivos)')
print(f'Sin código : {sin_codigo}')
print(f'Valor venta (activos con existencia): {valor:,.2f}')
print(f'Valor costo (activos con existencia): {costo:,.2f}')
if avisos:
    print(f'\nAvisos ({len(avisos)}):')
    for a in avisos:
        print('  -', a)
